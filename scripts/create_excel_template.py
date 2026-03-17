from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


def build_template(output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Roadmap"

    headers = ["application", "feature_name", "quarter", "month", "year"]
    sheet.append(headers)

    header_fill = PatternFill(start_color="1A5FFF", end_color="1A5FFF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 24,
        "B": 38,
        "C": 14,
        "D": 16,
        "E": 12,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width

    quarter_validation = DataValidation(type="list", formula1='"Q1,Q2,Q3,Q4"', allow_blank=True)
    month_validation = DataValidation(
        type="list",
        formula1='"January,February,March,April,May,June,July,August,September,October,November,December"',
        allow_blank=True,
    )
    year_validation = DataValidation(type="whole", operator="between", formula1="2020", formula2="2100", allow_blank=True)

    sheet.add_data_validation(quarter_validation)
    sheet.add_data_validation(month_validation)
    sheet.add_data_validation(year_validation)

    quarter_validation.add("C2:C1000")
    month_validation.add("D2:D1000")
    year_validation.add("E2:E1000")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:E1000"

    help_sheet = workbook.create_sheet(title="Instructions")
    help_sheet["A1"] = "Roadmap Template Instructions"
    help_sheet["A1"].font = Font(size=14, bold=True)
    help_sheet["A3"] = "Fill the Roadmap sheet with one roadmap item per row."
    help_sheet["A4"] = "Required columns: application, feature_name, quarter, month, year."
    help_sheet["A5"] = "Use quarter values: Q1, Q2, Q3, Q4."
    help_sheet["A6"] = "Use full month names (for example, January)."
    help_sheet["A7"] = "Use a 4-digit year (for example, 2026)."
    help_sheet.column_dimensions["A"].width = 95

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


if __name__ == "__main__":
    destination = Path("templates/roadmap_template.xlsx")
    build_template(destination)
    print(f"Created template at {destination}")
