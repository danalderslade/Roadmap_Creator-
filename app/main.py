from __future__ import annotations

import csv
import io
import re
from calendar import month_name
from datetime import datetime
from pathlib import Path
from typing import Any

from dateutil import parser as date_parser
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook

app = FastAPI(title="Roadmap Creator", version="1.0.0")

templates = Jinja2Templates(directory="app/templates")
app.mount("/static", StaticFiles(directory="app/static"), name="static")
EXCEL_TEMPLATE_PATH = Path("templates/roadmap_template.xlsx")


HEADER_SYNONYMS: dict[str, list[str]] = {
    "application": [
        "application",
        "app",
        "product",
        "project",
        "project key",
        "component",
        "component/s",
        "team",
    ],
    "feature_name": [
        "feature",
        "feature name",
        "summary",
        "title",
        "name",
        "epic name",
        "issue",
    ],
    "system": ["system", "platform", "service", "subsystem", "domain", "product line"],
    "region": ["region", "geo", "geography", "market", "country", "location"],
    "delivery_type": [
        "delivery type",
        "type",
        "work type",
        "classification",
        "initiative type",
        "deployment type",
    ],
    "quarter": ["quarter", "delivery quarter", "target quarter", "planned quarter", "qtr"],
    "month": ["month", "delivery month", "target month", "planned month"],
    "year": ["year", "delivery year", "target year", "planned year"],
    "date": [
        "due date",
        "target end",
        "target start",
        "release date",
        "planned date",
        "delivery date",
        "start date",
        "end date",
    ],
}


def normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.strip().lower()).strip()


def normalize_month(month_value: str) -> str | None:
    if not month_value:
        return None

    clean = month_value.strip()
    if not clean:
        return None

    if clean.isdigit():
        month_num = int(clean)
        if 1 <= month_num <= 12:
            return month_name[month_num]

    for i in range(1, 13):
        full = month_name[i]
        if clean.lower() in {full.lower(), full[:3].lower()}:
            return full

    return clean.title()


def quarter_from_month(month_number: int) -> str:
    return f"Q{((month_number - 1) // 3) + 1}"


def month_to_number(month_value: str) -> int | None:
    if not month_value:
        return None
    for i in range(1, 13):
        if month_value.lower() == month_name[i].lower():
            return i
    return None


def parse_year(value: str) -> str | None:
    if not value:
        return None
    stripped = value.strip()
    if re.fullmatch(r"\d{4}", stripped):
        return stripped
    try:
        parsed = date_parser.parse(stripped, fuzzy=True)
        return str(parsed.year)
    except (ValueError, TypeError, OverflowError):
        return None


def parse_date(value: str) -> datetime | None:
    if not value:
        return None
    try:
        return date_parser.parse(value, fuzzy=True, dayfirst=False)
    except (ValueError, TypeError, OverflowError):
        return None


def first_non_empty(row: dict[str, str], columns: list[str]) -> str | None:
    for col in columns:
        val = row.get(col, "")
        if val and val.strip():
            return val.strip()
    return None


def build_column_lookup(headers: list[str]) -> dict[str, list[str]]:
    normalized_map = {normalize_header(h): h for h in headers}
    result: dict[str, list[str]] = {k: [] for k in HEADER_SYNONYMS}

    for field, aliases in HEADER_SYNONYMS.items():
        for alias in aliases:
            normalized_alias = normalize_header(alias)
            if normalized_alias in normalized_map:
                result[field].append(normalized_map[normalized_alias])

    return result


def normalize_roadmap_rows(source_rows: list[dict[str, str]], headers: list[str]) -> list[dict[str, str]]:
    columns = build_column_lookup(headers)
    rows: list[dict[str, str]] = []

    for source_row in source_rows:
        application = first_non_empty(source_row, columns["application"]) or "Unspecified"
        feature_name = first_non_empty(source_row, columns["feature_name"]) or "Untitled feature"
        system = first_non_empty(source_row, columns["system"]) or application
        region = first_non_empty(source_row, columns["region"]) or "Global"
        delivery_type = first_non_empty(source_row, columns["delivery_type"]) or "Feature"
        quarter = first_non_empty(source_row, columns["quarter"]) or ""
        month = first_non_empty(source_row, columns["month"]) or ""
        year = first_non_empty(source_row, columns["year"]) or ""

        parsed_date = None
        if not month or not year or not quarter:
            date_candidate = first_non_empty(source_row, columns["date"]) or ""
            parsed_date = parse_date(date_candidate)

        normalized_month = normalize_month(month) if month else None

        if not normalized_month and parsed_date:
            normalized_month = month_name[parsed_date.month]

        normalized_year = parse_year(year) if year else None
        if not normalized_year and parsed_date:
            normalized_year = str(parsed_date.year)

        normalized_quarter = quarter.strip().upper() if quarter else ""
        if normalized_quarter and not normalized_quarter.startswith("Q") and normalized_quarter.isdigit():
            normalized_quarter = f"Q{normalized_quarter}"
        if not normalized_quarter and parsed_date:
            normalized_quarter = quarter_from_month(parsed_date.month)
        if not normalized_quarter and normalized_month:
            month_index = month_to_number(normalized_month)
            if month_index:
                normalized_quarter = quarter_from_month(month_index)

        rows.append(
            {
                "application": application,
                "feature_name": feature_name,
                "system": system,
                "region": region,
                "delivery_type": delivery_type.title(),
                "quarter": normalized_quarter or "Unspecified",
                "month": normalized_month or "Unspecified",
                "year": normalized_year or "Unspecified",
            }
        )

    return rows


def extract_roadmap_rows_from_csv(csv_text: str) -> list[dict[str, str]]:
    reader = csv.DictReader(io.StringIO(csv_text))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="The uploaded CSV has no header row.")
    return normalize_roadmap_rows(list(reader), list(reader.fieldnames))


def extract_roadmap_rows_from_excel(content: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    worksheet = workbook.active
    row_iter = worksheet.iter_rows(values_only=True)

    try:
        raw_headers = next(row_iter)
    except StopIteration as exc:
        raise HTTPException(status_code=400, detail="The uploaded Excel file is empty.") from exc

    headers = [str(cell).strip() if cell is not None else "" for cell in raw_headers]
    if not any(headers):
        raise HTTPException(status_code=400, detail="The uploaded Excel file has no header row.")

    source_rows: list[dict[str, str]] = []
    for raw_row in row_iter:
        values = ["" if cell is None else str(cell).strip() for cell in raw_row]
        if not any(values):
            continue
        row_dict = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        source_rows.append(row_dict)

    return normalize_roadmap_rows(source_rows, headers)


def build_filters(rows: list[dict[str, str]]) -> dict[str, list[str]]:
    keys = ["application", "feature_name", "system", "region", "delivery_type", "quarter", "month", "year"]
    filters: dict[str, list[str]] = {}
    for key in keys:
        values = sorted({row[key] for row in rows}, key=lambda x: (x == "Unspecified", x))
        filters[key] = values
    return filters


@app.get("/", response_class=HTMLResponse)
async def home(request: Request) -> HTMLResponse:
    return templates.TemplateResponse("index.html", {"request": request})


@app.post("/api/upload")
async def upload_csv(file: UploadFile = File(...)) -> dict[str, Any]:
    lower_name = file.filename.lower()
    if not (lower_name.endswith(".csv") or lower_name.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="Please upload a CSV or XLSX file.")

    content = await file.read()
    if lower_name.endswith(".xlsx"):
        rows = extract_roadmap_rows_from_excel(content)
    else:
        decoded = None
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                decoded = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue

        if decoded is None:
            raise HTTPException(status_code=400, detail="Could not decode file. Please upload UTF-8 CSV.")

        rows = extract_roadmap_rows_from_csv(decoded)

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows were found in the file.")

    return {
        "total_rows": len(rows),
        "rows": rows,
        "filters": build_filters(rows),
    }


@app.get("/healthz")
async def healthz() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/template")
async def download_template() -> FileResponse:
    if not EXCEL_TEMPLATE_PATH.exists():
        raise HTTPException(status_code=404, detail="Excel template not found.")

    return FileResponse(
        path=EXCEL_TEMPLATE_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="roadmap_template.xlsx",
    )
